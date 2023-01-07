from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from random_user_agent.user_agent import UserAgent
from random_user_agent.params import SoftwareName, OperatingSystem
import time
import random
import calendar
import cfscrape
import os


scraper = cfscrape.create_scraper() #ANTI-CDN

# random UserAgent Setting
software_names = [SoftwareName.CHROME.value]
operating_systems = [OperatingSystem.WINDOWS.value, OperatingSystem.LINUX.value]   
user_agent_rotator = UserAgent(software_names=software_names, operating_systems=operating_systems, limit=100)

def rq(url):
    """
    連結獲取 html
    """
    headers = {'User-Agent': user_agent_rotator.get_random_user_agent()}
    proxy_ips = ['51.15.227.220:3128', '81.162.56.154:8081', '133.18.195.135:8080', '66.29.154.103:3128', '164.132.170.100:80', '47.241.245.186:80', '80.48.119.28:8080', '3.212.9.208:80', '51.15.242.202:8888']
    ip = random.choice(proxy_ips)
    resp = scraper.get(url=url,
                        proxies={'http': 'http://' + ip}, headers=headers)
    hf = BeautifulSoup(resp.text, "html.parser")
    return hf


search = input("請輸入關鍵字: ")
os.system('clear')
pa_name = input('請輸入要存檔的檔案名稱: ') + ".xlsx"
os.system('clear')
lll_num = int(input("請輸入流水號開始號碼: "))
os.system('clear')
save_row = int(input("請輸入開始行數: "))
os.system('clear')

print("建檔模式代碼:\n1 一次完成\n2 逐頁確認")
save_mode = int(input('請輸入建檔模式： '))
os.system('clear')

print("搜尋模式:\n1 常規搜尋\n2 日期搜尋")
mode = int(input("請選擇模式: "))
os.system('clear')

if mode == 2:
    print("請選擇日期範圍: ")
    print("1. 今天")
    print("2. 過去 7 天")
    print("3. 過去 30 天")
    print("4. 過去 12 個月")
    print("5. 過去 2 年")
    print("6. 過去 5 年")
    date_mode = int(input("請輸入代碼"))
    if date_mode == 1:
        url = "https://www.nature.com/search?q="+str(search)+"&date_range=today&order=relevance"
    elif date_mode == 2:
        url = "https://www.nature.com/search?q="+str(search)+"&date_range=last_7_days&order=relevance"
    elif date_mode == 3:
        url = "https://www.nature.com/search?q="+str(search)+"&date_range=last_30_days&order=relevance"
    elif date_mode == 4:
        url = "https://www.nature.com/search?q="+str(search)+"&date_range=last_year&order=relevance"
    elif date_mode == 5:
        url = "https://www.nature.com/search?q="+str(search)+"&date_range=last_2_years&order=relevance"
    elif date_mode == 6:
        url = "https://www.nature.com/search?q="+str(search)+"&date_range=last_5_years&order=relevance"
    os.system('clear')
    
else:
    url = "https://www.nature.com/search?q="+str(search)+"&order=relevance"
    
hfe = rq(url)

page = hfe.select('#content > div > div > nav > ul > li:nth-child(6) > a')

for i in page:
    page_all = i.get_text()
    page_all = page_all.split(" ")
    page_all = page_all[1]
    page_all = int(page_all)

print("總共有", page_all, "頁")
while True:
    page_num = int(input("請輸入要搜尋到第幾頁: "))
    if (page_num > page_all):
        print("已經超過頁數上限，請重新輸入")
    else:
        page_all = page_num
        break

os.system('clear')

author_name_column = 2
author_mail_column = 3
author_paper_column = 4
author_date_column = 5
author_lll_column = 1

for page in range(page_all):
    os.system('clear')
    print("當前進度為第", page+1 ,"/", page_all, "頁")
    if save_mode == 2:
        if input('想停止請輸入 END,若想繼續請按 Enter: ') == "END":
            break

    url = url+"&page="+str(page+1)
    
    hf = rq(url)

    title = hf.select('#search-article-list > div > ul > li > div > article > div.c-card__layout.u-full-height > div.c-card__body.u-display-flex.u-flex-direction-column > h3 > a')

    for item in title:      

        try:
            wb = load_workbook(pa_name)
            sheet = wb['1']
            
        except:
            print("讀取不到 excel，將創建新的 excel")
            wb = Workbook()
            wsl = wb.active
            wsl.title = '1'
            wb.save(pa_name)
            wb = load_workbook(pa_name)
            sheet = wb['1']
        
        wb.iso_dates = True
        
        
        author_url = 'https://www.nature.com'+item.get('href')
        
        paper = rq(author_url)

        for t in paper.select('#content > main > article > div.c-article-header > header > ul.c-article-identifiers > li:nth-child(2) > a > time'):
            date_value= t.get_text()
            date_value = date_value.split(' ')
            date_value[1] = list(calendar.month_name).index(date_value[1])
            
        paper_name = paper.select('#content > main > article > div.c-article-header > header > h1')
        for i in paper_name:
            paper_name = i.get_text()

        author = paper.select('#corresponding-author-list > a')

        for i in author:
            name = i.get_text()
            mail = i.get('href')
            mail = mail.split(':')
            mail = mail[1]
            name_save = sheet.cell(row= save_row, column=author_name_column)
            mail_save = sheet.cell(row= save_row, column=author_mail_column)
            paper_save = sheet.cell(row= save_row, column=author_paper_column)
            date_save = sheet.cell(row= save_row, column=author_date_column)
            lll_save = sheet.cell(row = save_row, column= author_lll_column)
            try:
                year = format(date_value[2])
                month = format(date_value[1])
                day = format(date_value[0])
                date_save.value = year+"/"+month+"/"+day
            except:
                print("無法找到日期")


            paper_save.value = paper_name
            paper_save.hyperlink = author_url
            name_save.value = name
            mail_save.value = mail
            lll_save.value = lll_num

            save_row += 1
            print('.', end='', flush=True)          
            
        lll_num += 1
        time.sleep(random.randint(1,3))
        wb.save(pa_name)

print(".")
print("---------------------------")
print('已完成')
        
    
        


